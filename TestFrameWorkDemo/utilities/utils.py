import softest
from openpyxl import Workbook, load_workbook
class utils(softest.TestCase):
    def assertListText(self,list,value):
        for i in list:
            print(i.text)
            # assert i.text == value
            # print("assert pass")
            self.soft_assert(self.assertEqual, i.text, value)
            if i.text == value:
                print("assert pass")
            else:
                print("assert fail")
        self.assert_all()
    def read_data_from_excel(self,filename,sheetname):
        datalist=[]
        wb= load_workbook(filename=filename)
        sh=wb[sheetname]
        row_ct=sh.max_row
        col_ct=sh.max_column
        for i in range(2,row_ct+1):
            row=[]
            for j in range(1,col_ct+1):
                row.append(sh.cell(row=i,column=j).value)
            datalist.append(row)
        return datalist
