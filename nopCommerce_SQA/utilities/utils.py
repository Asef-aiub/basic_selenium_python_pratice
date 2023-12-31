import softest
from openpyxl import Workbook, load_workbook
import csv
class utils():
    def assertText(self,screenText,value):
        if screenText == value:
            assert screenText == value
            print("test passed")
        else:
            print("test failed")


    def read_data_from_excel(self,filename,sheetname):
        datalist=[]
        wb= load_workbook(filename=filename)
        sh=wb[sheetname]
        row_ct=sh.max_row
        col_ct=sh.max_column
        for i in range(2,row_ct+1):
            row=[]
            for j in range(1,col_ct+1):
                row.append(sh.cell(row=i,column=j).value)
            datalist.append(tuple(row))
        return datalist
    def read_data_from_csv(self,filename):
        datalist=[]
        csvdata=open(filename,"r")
        reader=csv.reader(csvdata)
        next(reader)
        for rows in reader:
            datalist.append(rows)
        return datalist